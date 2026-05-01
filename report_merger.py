"""
Report Merger Module

Merges two MATCH_REPORT .xlsx files (or their _data.pkl counterparts) produced
by PopulationMatcher into a single deduplicated MERGED_MATCH_REPORT.xlsx.

Merge logic
-----------
Matched rows (match_type != 'NO_MATCH'):
  Group by (ref_SSN, filename). Same participant on same page across both files
  is a duplicate — keep one row per group using:
    1. Higher-priority match_type per MATCH_HIERARCHY
    2. Tie → most non-empty extracted_* columns
    3. Tie → file1
  Rows with different ref_SSN on the same filename are different participants —
  kept independently. Rows unique to file2 (new docids) pass through untouched.

NO_MATCH rows (match_type == 'NO_MATCH'):
  Group by filename. For each cross-file row pair, check for any shared
  (element, extracted value). If overlap exists → same underlying extraction →
  keep the more populated row (tie → file1). No overlap → different extractions
  → keep both. File-only rows pass through untouched.

Post-merge:
  page_number is re-derived from filename using the pattern {docid}_page{n}.pdf,
  overwriting any broken values from either input file.

Usage
-----
  python report_merger.py file1.xlsx file2.xlsx [--output-dir ./output] [--name merged]

  Or import and call merge_reports() directly.
"""

import re
import sys
import logging
import argparse
import pickle
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd

logger = logging.getLogger(__name__)

MATCH_HIERARCHY = [
    'FULL_SSN',
    'FULL_SSN_DISCREPANCY',
    'FULL_LAST6',
    'FULL_LAST4',
    'FULL_DOB_DOH',
    'FNAME_LNAME_DOB',
    'FULL_LNAME_DOB',
    'FNAME_DOB_LI',
    'LNAME_DOB_MY',
    'OTHER_SSN',
    'FULL_NAME',
    'PARTIAL',
    'NO_MATCH',
]

HIERARCHY_RANK: Dict[str, int] = {mt: i for i, mt in enumerate(MATCH_HIERARCHY)}


# ------------------------------------------------------------------ #
#  Helpers                                                             #
# ------------------------------------------------------------------ #

def _extracted_cols(df: pd.DataFrame) -> List[str]:
    return [c for c in df.columns if c.startswith('extracted_')]


def _populated_count(row: pd.Series, ext_cols: List[str]) -> int:
    """Number of non-empty extracted_* values in a row."""
    return sum(1 for c in ext_cols if str(row.get(c, '')).strip())


def _rank(match_type: str) -> int:
    return HIERARCHY_RANK.get(str(match_type), 999)


def _page_number_from_filename(filename: str) -> Optional[int]:
    """Extract page number from '{docid}_page{n}.pdf' filename."""
    m = re.search(r'[_\-][Pp]age[_\-]?(\d+)', str(filename), re.IGNORECASE)
    if not m:
        # also try bare _pageN without separator before digit
        m = re.search(r'page(\d+)', str(filename), re.IGNORECASE)
    return int(m.group(1)) if m else None


def _load_file(path: str) -> pd.DataFrame:
    """Load a MATCH_REPORT from .xlsx or _data.pkl."""
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"File not found: {path}")
    if p.suffix == '.pkl' or path.endswith('_data.pkl'):
        with open(p, 'rb') as f:
            df = pickle.load(f)
        if not isinstance(df, pd.DataFrame):
            raise ValueError(f"Pickle at {path} did not contain a DataFrame")
        return df
    # xlsx — read first sheet that isn't 'Stats'
    xl = pd.ExcelFile(path)
    frames = []
    for sheet in xl.sheet_names:
        if sheet.lower() == 'stats':
            continue
        frames.append(pd.read_excel(xl, sheet_name=sheet, dtype=str).fillna(''))
    if not frames:
        raise ValueError(f"No data sheets found in {path}")
    return pd.concat(frames, ignore_index=True)


# ------------------------------------------------------------------ #
#  Core merge steps                                                    #
# ------------------------------------------------------------------ #

def _resolve_matched(matched: pd.DataFrame, ext_cols: List[str]) -> pd.DataFrame:
    """
    Deduplicate matched rows.

    Group by (ref_SSN, filename). Within each group keep the single best row:
      1. Lowest HIERARCHY_RANK (best match type)
      2. Most populated extracted_* columns
      3. file1 preferred (source == 'file1' sorts first)
    """
    if matched.empty:
        return matched

    def best_row(group: pd.DataFrame) -> pd.Series:
        if len(group) == 1:
            return group.iloc[0]
        ranked = group.copy()
        ranked['_rank'] = ranked['match_type'].map(_rank)
        ranked['_pop']  = ranked.apply(lambda r: _populated_count(r, ext_cols), axis=1)
        # _source: file1 < file2 → sort ascending keeps file1 first on tie
        ranked['_src_order'] = ranked['_source'].map({'file1': 0, 'file2': 1}).fillna(1)
        ranked = ranked.sort_values(['_rank', '_pop', '_src_order'],
                                    ascending=[True, False, True])
        return ranked.iloc[0].drop(labels=['_rank', '_pop', '_src_order'])

    result = (
        matched
        .groupby(['ref_SSN', 'filename'], sort=False, group_keys=False)
        .apply(best_row)
        .reset_index(drop=True)
    )
    return result


def _resolve_nomatch(nomatch: pd.DataFrame, ext_cols: List[str]) -> pd.DataFrame:
    """
    Deduplicate NO_MATCH rows per filename.

    For each filename, compare every file1 row against every file2 row.
    If a pair shares at least one (element, extracted_value) → same extraction →
    keep the more populated row (tie → file1), discard the other.
    Rows with no cross-file overlap on any pair → kept as-is.
    Rows that only appear in one file → kept as-is.
    """
    if nomatch.empty:
        return nomatch

    kept_rows: List[pd.Series] = []

    for filename, group in nomatch.groupby('filename', sort=False):
        file1_rows = group[group['_source'] == 'file1'].reset_index(drop=True)
        file2_rows = group[group['_source'] == 'file2'].reset_index(drop=True)

        if file1_rows.empty:
            # All from file2 — pass through
            for _, r in file2_rows.iterrows():
                kept_rows.append(r)
            continue
        if file2_rows.empty:
            # All from file1 — pass through
            for _, r in file1_rows.iterrows():
                kept_rows.append(r)
            continue

        # Build extracted value sets per row: {elem: value} where value != ''
        def elem_value_set(row: pd.Series) -> set:
            return {
                (col, str(row[col]).strip())
                for col in ext_cols
                if str(row.get(col, '')).strip()
            }

        f1_sets = [elem_value_set(row) for _, row in file1_rows.iterrows()]
        f2_sets = [elem_value_set(row) for _, row in file2_rows.iterrows()]

        # Track which rows are consumed by a matched pair
        f1_consumed = [False] * len(file1_rows)
        f2_consumed = [False] * len(file2_rows)

        for i, f1_row in file1_rows.iterrows():
            for j, f2_row in file2_rows.iterrows():
                if f1_consumed[i] or f2_consumed[j]:
                    continue
                overlap = f1_sets[i] & f2_sets[j]
                if not overlap:
                    continue
                # Same extraction — pick more populated, tie → file1
                f1_pop = _populated_count(f1_row, ext_cols)
                f2_pop = _populated_count(f2_row, ext_cols)
                winner = f1_row if f1_pop >= f2_pop else f2_row
                kept_rows.append(winner)
                f1_consumed[i] = True
                f2_consumed[j] = True

        # Keep unconsumed rows from both files
        for i, f1_row in file1_rows.iterrows():
            if not f1_consumed[i]:
                kept_rows.append(f1_row)
        for j, f2_row in file2_rows.iterrows():
            if not f2_consumed[j]:
                kept_rows.append(f2_row)

    return pd.DataFrame(kept_rows).reset_index(drop=True)


# ------------------------------------------------------------------ #
#  Output                                                              #
# ------------------------------------------------------------------ #

def _save_merged_report(df: pd.DataFrame, output_dir: Path, name: str) -> str:
    """Save merged DataFrame as MERGED_MATCH_REPORT.xlsx with per-docid tabs."""
    from openpyxl.styles import PatternFill, Font

    def fill(hex_color: str) -> PatternFill:
        return PatternFill(fill_type='solid', fgColor=hex_color)

    EXTRACTED_COL    = fill('EEF4FF')
    EXTRACTED_HEADER = fill('BDD7EE')
    REF_COL          = fill('E8E8E8')
    REF_HEADER       = fill('C0C0C0')

    MATCH_TYPE_FILL: Dict[str, PatternFill] = {
        'FULL_SSN':             fill('92D050'),
        'FULL_SSN_DISCREPANCY': fill('FFC000'),
        'FULL_LAST6':           fill('00B050'),
        'FULL_LAST4':           fill('70AD47'),
        'FULL_DOB_DOH':         fill('A9D18E'),
        'FNAME_LNAME_DOB':      fill('C6EFCE'),
        'FULL_LNAME_DOB':       fill('E2EFDA'),
        'FNAME_DOB_LI':         fill('FFF2CC'),
        'LNAME_DOB_MY':         fill('FFE699'),
        'OTHER_SSN':            fill('F4B183'),
        'FULL_NAME':            fill('9DC3E6'),
        'PARTIAL':              fill('FFEB9C'),
        'NO_MATCH':             fill('FFC7CE'),
    }

    SOURCE_FILL: Dict[str, PatternFill] = {
        'census_1':        fill('BDD7EE'),
        'census_2':        fill('9BC2E6'),
        'metadata':        fill('E2D0F0'),
        'metadata/census': fill('C5D4EB'),
        'other_ssn':       fill('F4B183'),
    }

    DISCREPANCY_FILL = fill('FF9999')

    # Sort by docid → hierarchy → page_number
    df = df.copy()
    df['_sort_key'] = df['match_type'].map(lambda x: HIERARCHY_RANK.get(x, 999))
    if 'page_number' in df.columns:
        df['page_number'] = pd.to_numeric(df['page_number'], errors='coerce').fillna(0).astype(int)
    df = df.sort_values(
        ['docid', '_sort_key', 'page_number'] if 'page_number' in df.columns
        else ['docid', '_sort_key']
    ).drop(columns=['_sort_key'])

    # Assign sheet names (one per docid)
    unique_docids = df['docid'].unique()
    docid_to_sheet: Dict[str, str] = {}
    used_sheets: set = set()
    unknown_counter = 1

    for docid in unique_docids:
        docid_str = str(docid).strip()
        if not docid_str:
            sheet_name = f"Unknown_{unknown_counter}"
            unknown_counter += 1
        else:
            base = docid_str[:31]
            if base not in used_sheets:
                sheet_name = base
            else:
                counter = 2
                while f"{base[:28]}_{counter}" in used_sheets:
                    counter += 1
                sheet_name = f"{base[:28]}_{counter}"
        docid_to_sheet[docid] = sheet_name
        used_sheets.add(sheet_name)

    df['_sheet'] = df['docid'].map(docid_to_sheet)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"{name}_MERGED_MATCH_REPORT_{timestamp}.xlsx"
    output_dir.mkdir(parents=True, exist_ok=True)

    bold_font  = Font(bold=True)
    link_font  = Font(color='0563C1', underline='single')

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # ---- Stats sheet ------------------------------------------------
        hierarchy_order = HIERARCHY_RANK
        no_match_rank   = hierarchy_order.get('NO_MATCH', 999)
        stats_rows = []

        for docid_val, docid_df in df.groupby('docid', sort=False):
            stat: Dict = {'DOCID': docid_val}
            stat['Unique Pages'] = (
                docid_df['filename'].nunique() if 'filename' in docid_df.columns else 0
            )
            if 'ref_SSN' in docid_df.columns:
                non_empty_mask = docid_df['ref_SSN'].astype(str).str.strip() != ''
                non_empty_ssn  = docid_df.loc[non_empty_mask, 'ref_SSN']
                stat['Unique Participants'] = non_empty_ssn.nunique()
                stat['Total Participants']  = int(non_empty_mask.sum())
            else:
                stat['Unique Participants'] = 0
                stat['Total Participants']  = 0

            if 'filename' in docid_df.columns and 'extraction_order' in docid_df.columns:
                pair_best = (
                    docid_df
                    .assign(_rank=docid_df['match_type'].map(
                        lambda x: hierarchy_order.get(x, 999)))
                    .groupby(['filename', 'extraction_order'])['_rank']
                    .min()
                )
                stat['Unmatched'] = int((pair_best == no_match_rank).sum())
            else:
                stat['Unmatched'] = 0

            src = docid_df['matched_source'].astype(str)
            stat['Census 1']        = int((src == 'census_1').sum())
            stat['Census 2']        = int((src == 'census_2').sum())
            stat['Metadata']        = int((src == 'metadata').sum())
            stat['Metadata/Census'] = int((src == 'metadata/census').sum())
            stat['Other SSN']       = int(src.str.lower().str.startswith('other ssn:').sum())
            stats_rows.append(stat)

        stats_df = pd.DataFrame(stats_rows)
        stats_df.to_excel(writer, sheet_name='Stats', index=False)
        ws_stats = writer.sheets['Stats']
        for col_idx in range(1, len(stats_df.columns) + 1):
            cell = ws_stats.cell(row=1, column=col_idx)
            cell.font = bold_font
            cell.fill = fill('D9D9D9')
        for col_idx in range(1, len(stats_df.columns) + 1):
            col_letter = ws_stats.cell(1, col_idx).column_letter
            max_len = max(
                (len(str(ws_stats.cell(r, col_idx).value or ''))
                 for r in range(1, len(stats_df) + 2)),
                default=0,
            )
            ws_stats.column_dimensions[col_letter].width = min(max_len + 2, 40)

        # ---- Data sheets (one per docid) ------------------------------------
        for sheet_name, sheet_df in df.groupby('_sheet', sort=False):
            sheet_df = sheet_df.drop(columns=['_sheet']).reset_index(drop=True)

            pdf_urls = sheet_df['_pdf_url'].tolist() if '_pdf_url' in sheet_df.columns else []
            if '_pdf_url' in sheet_df.columns:
                sheet_df = sheet_df.drop(columns=['_pdf_url'])

            # Drop entirely-empty columns
            non_empty_cols = [
                c for c in sheet_df.columns
                if sheet_df[c].replace('', pd.NA).notna().any()
            ]
            sheet_df = sheet_df[non_empty_cols]

            sheet_df.to_excel(writer, sheet_name=str(sheet_name), index=False)
            ws   = writer.sheets[str(sheet_name)]
            cols = list(sheet_df.columns)

            # Hyperlinks on filename column
            if 'filename' in cols and pdf_urls:
                fn_col_idx = cols.index('filename') + 1
                for row_idx, url in enumerate(pdf_urls, start=2):
                    if url:
                        cell = ws.cell(row=row_idx, column=fn_col_idx)
                        cell.hyperlink = url
                        cell.font = link_font

            for col_idx, col_name in enumerate(cols, start=1):
                header_cell = ws.cell(row=1, column=col_idx)

                if col_name.startswith('extracted_'):
                    header_cell.fill = EXTRACTED_HEADER
                    for row_idx in range(2, len(sheet_df) + 2):
                        ws.cell(row=row_idx, column=col_idx).fill = EXTRACTED_COL
                elif col_name.startswith('ref_'):
                    header_cell.fill = REF_HEADER
                    for row_idx in range(2, len(sheet_df) + 2):
                        ws.cell(row=row_idx, column=col_idx).fill = REF_COL

                if col_name == 'match_type':
                    for row_idx in range(2, len(sheet_df) + 2):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell_fill = MATCH_TYPE_FILL.get(str(cell.value or ''))
                        if cell_fill:
                            cell.fill = cell_fill
                elif col_name == 'matched_source':
                    for row_idx in range(2, len(sheet_df) + 2):
                        cell     = ws.cell(row=row_idx, column=col_idx)
                        cell_val = str(cell.value or '')
                        if cell_val.lower().startswith('other ssn:'):
                            cell.fill = SOURCE_FILL['other_ssn']
                        else:
                            src_fill = SOURCE_FILL.get(cell_val.lower())
                            if src_fill:
                                cell.fill = src_fill

                max_len = max(
                    (len(str(cell.value or '')) for cell in ws[ws.cell(1, col_idx).column_letter]),
                    default=0,
                )
                ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = min(max_len + 2, 50)

            # Discrepancy highlighting
            if 'discrepancy_fields' in cols:
                disc_col_idx = cols.index('discrepancy_fields') + 1
                for row_idx in range(2, len(sheet_df) + 2):
                    disc_value = str(ws.cell(row=row_idx, column=disc_col_idx).value or '')
                    if disc_value:
                        for elem in [e.strip() for e in disc_value.split(',') if e.strip()]:
                            for prefix in ('extracted_', 'ref_'):
                                target_col = f'{prefix}{elem}'
                                if target_col in cols:
                                    ws.cell(row=row_idx, column=cols.index(target_col) + 1).fill = DISCREPANCY_FILL

    logger.info(f"Merged report saved: {output_path}")
    print(f"  Saved: {output_path}")
    return str(output_path)


# ------------------------------------------------------------------ #
#  Public entry point                                                  #
# ------------------------------------------------------------------ #

def merge_reports(
    file1_path: str,
    file2_path: str,
    output_dir: str = '.',
    name: str = 'merged',
) -> str:
    """
    Merge two MATCH_REPORT files into a single deduplicated report.

    Args:
        file1_path:  Path to first MATCH_REPORT (.xlsx or _data.pkl).
        file2_path:  Path to second MATCH_REPORT (.xlsx or _data.pkl).
        output_dir:  Directory to write the merged report into.
        name:        Prefix for the output filename.

    Returns:
        Path to the saved MERGED_MATCH_REPORT.xlsx.
    """
    print(f"\nLoading file 1: {file1_path}")
    df1 = _load_file(file1_path)
    df1['_source'] = 'file1'
    print(f"  {len(df1)} rows loaded")

    print(f"Loading file 2: {file2_path}")
    df2 = _load_file(file2_path)
    df2['_source'] = 'file2'
    print(f"  {len(df2)} rows loaded")

    # Align columns — fill missing columns with empty string
    all_cols = list(dict.fromkeys(list(df1.columns) + list(df2.columns)))
    for col in all_cols:
        if col not in df1.columns:
            df1[col] = ''
        if col not in df2.columns:
            df2[col] = ''

    combined = pd.concat([df1[all_cols], df2[all_cols]], ignore_index=True)
    ext_cols  = _extracted_cols(combined)

    print(f"\nTotal rows before merge: {len(combined)}")

    # Split matched vs no_match
    is_nomatch  = combined['match_type'].astype(str).str.strip() == 'NO_MATCH'
    matched_df  = combined[~is_nomatch].copy()
    nomatch_df  = combined[is_nomatch].copy()

    print(f"  Matched rows:   {len(matched_df)}")
    print(f"  NO_MATCH rows:  {len(nomatch_df)}")

    # Resolve each group
    resolved_matched = _resolve_matched(matched_df, ext_cols)
    resolved_nomatch = _resolve_nomatch(nomatch_df, ext_cols)

    print(f"\nAfter deduplication:")
    print(f"  Matched rows:   {len(resolved_matched)}")
    print(f"  NO_MATCH rows:  {len(resolved_nomatch)}")

    # Recombine
    final = pd.concat([resolved_matched, resolved_nomatch], ignore_index=True)

    # Drop internal tracking column
    if '_source' in final.columns:
        final = final.drop(columns=['_source'])

    # Re-derive page_number from filename
    if 'filename' in final.columns:
        final['page_number'] = final['filename'].apply(
            lambda fn: _page_number_from_filename(fn) or 0
        )

    print(f"  Total rows in merged report: {len(final)}")

    output_path = _save_merged_report(final, Path(output_dir), name)
    return output_path


# ------------------------------------------------------------------ #
#  CLI                                                                 #
# ------------------------------------------------------------------ #

def main():
    logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')

    parser = argparse.ArgumentParser(
        description='Merge two MATCH_REPORT files into a deduplicated report.'
    )
    parser.add_argument('file1', help='Path to first MATCH_REPORT (.xlsx or _data.pkl)')
    parser.add_argument('file2', help='Path to second MATCH_REPORT (.xlsx or _data.pkl)')
    parser.add_argument('--output-dir', default='.', help='Output directory (default: current dir)')
    parser.add_argument('--name', default='merged', help='Output filename prefix (default: merged)')
    args = parser.parse_args()

    result = merge_reports(
        file1_path=args.file1,
        file2_path=args.file2,
        output_dir=args.output_dir,
        name=args.name,
    )

    if result:
        print(f"\nDone. Merged report: {result}")
        sys.exit(0)
    else:
        print("\nMerge failed — check logs above.")
        sys.exit(1)


if __name__ == '__main__':
    main()
